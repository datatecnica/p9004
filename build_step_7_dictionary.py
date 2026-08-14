"""Step 7 — assemble the release dataset and its data dictionary.

Concatenates merged + derived + PCs, then builds a dictionary that is asserted to be in
strict 1:1 correspondence with the dataset columns.

Dictionary sources, in priority order per column:
  1. the curated cut's own dictionary          (clinical, 212 entries — authoritative)
  2. generated entries                         (p9001, key, PCs, corrected derivations)
  3. the previous release's dictionary         (proteomics, plate IDs, join artifacts,
                                                and the derived variables unchanged here)

Outputs:
  Project_9004_Unified_Emerging_Biomarkers.tab
  Project_9004_Data_Dictionary.tab
  Project_9004_Data_Dictionary.xlsx     (one row per entry; see note at the bottom)
"""

from __future__ import annotations

import glob
import os
import sys

import pandas as pd

from build_common import TIMESTAMP, log

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "build_intermediates")
PREV_DICT = os.path.join(HERE, os.pardir, "Project_9004_Data_Dictionary.tab")

DICT_COLS = ["Category", "Variable", "Description", "Code", "Decode",
             "Derived Variable", "Original Variable(s)", "Original Dataset(s)",
             "Derivation Notes"]

PC_BLOCK_DESC = {
    "p277_CSF": "Project 277 Olink Explore HT CSF",
    "p282_CNS_CSF": "Project 282 NULISA CNS Disease CSF",
    "p282_Inflammation_CSF": "Project 282 NULISA Inflammation CSF",
    "p288_CNS_plasma": "Project 288 NULISA CNS Disease plasma",
    "p288_Inflammation_plasma": "Project 288 NULISA Inflammation plasma",
    "p293_olink_plasma": "Project 293 Olink Explore HT plasma",
    "p312_Inflammation_CSF": "Project 312 NULISA Inflammation CSF",
    "p312_Inflammation_Plasma": "Project 312 NULISA Inflammation plasma",
    "p312_Neuro_CSF": "Project 312 NULISA Neuro 220 CSF",
    "p312_Neuro_Plasma": "Project 312 NULISA Neuro 220 plasma",
    "p314_CSF": "Project 314 Olink Explore HT CSF",
    "p314_Plasma": "Project 314 Olink Explore HT plasma",
}


def latest(pattern: str) -> str:
    hits = sorted(glob.glob(os.path.join(OUT, pattern)))
    if not hits:
        sys.exit(f"missing {pattern} — run the prior step first")
    return hits[-1]


def entry(**kw) -> dict:
    return {
        "Category": kw.get("category", "Clinical"),
        "Variable": kw["variable"],
        "Description": kw["description"],
        "Code": kw.get("code", "-"),
        "Decode": kw.get("decode", "-"),
        "Derived Variable": kw.get("derived", "Yes"),
        "Original Variable(s)": kw.get("original", "-"),
        "Original Dataset(s)": kw.get("dataset", "-"),
        "Derivation Notes": kw.get("notes", ""),
    }


def build_p9001_entries(cols: list[str]) -> list[dict]:
    """Entries for the new GP2 release-12 genetics block."""
    out = []
    n_snps = sum(1 for c in cols if c.startswith("p9001_rs"))
    for c in cols:
        stem = c[len("p9001_"):]
        if stem.startswith("Genetic_PRS_PRS"):
            n = stem.replace("Genetic_PRS_PRS", "")
            out.append(entry(
                category="Genetics", variable=c,
                description=f"Parkinson's disease polygenic risk score, GP2 release 12 (variant set {n})",
                original=stem, dataset="GP2_R12_PPMI_PRS_PCs.csv",
                notes=("One of four PRS variants shipped in the GP2 release-12 file (PRS157, "
                       f"PRS154, PRS152, PRS149); the file carries {n_snps} rs-dosage columns. "
                       "The variant sets behind each suffix are described in detail in the "
                       "Project 9001 documentation — consult it before choosing between them, "
                       "as they are not interchangeable.")))
        elif stem.startswith("Genetic_PRS_PC"):
            n = stem.replace("Genetic_PRS_PC", "")
            out.append(entry(
                category="Genetics", variable=c,
                description=f"Within-population genetic principal component {n}, GP2 release 12",
                original=stem, dataset="GP2_R12_PPMI_PRS_PCs.csv",
                notes=("Ancestry principal component computed within the inferred population. "
                       "Distinct from GP2_PC1-10, which come from the earlier GP2 release and "
                       "are retained alongside these.")))
        elif stem == "Genetic_PRS_InfPop":
            out.append(entry(
                category="Genetics", variable=c,
                description="Genetically inferred ancestry population, GP2 release 12",
                code="-", decode="EUR; AJ; AMR; AFR; AAC; EAS; MDE; CAH; CAS; SAS; FIN",
                original=stem, dataset="GP2_R12_PPMI_PRS_PCs.csv",
                notes=("Counterpart to GP2_nba_label from the earlier release, over a larger "
                       "participant set. The two are retained side by side and may disagree; "
                       "analysis strata are still defined on GP2_nba_label.")))
        else:  # rs dosage
            rsid, _, allele = stem.rpartition("_")
            out.append(entry(
                category="Genetics", variable=c, derived="No",
                description=f"Genotype dosage of the {allele} allele at {rsid}",
                code="0; 1; 2", decode="0 copies; 1 copy; 2 copies",
                original=stem, dataset="GP2_R12_PPMI_PRS_PCs.csv",
                notes="Parkinson's disease GWAS risk locus. Additive dosage of the named allele."))
    return out


def build_pc_entries(variance: pd.DataFrame) -> list[dict]:
    out = []
    for _, r in variance.iterrows():
        block = r["block"]
        desc = PC_BLOCK_DESC.get(block)
        if desc is None and block.startswith("harmonized_"):
            b = block[len("harmonized_"):]
            plat, fluid = BLOCK_DESC.get(b, (b, ""))
            desc = f"harmonized {plat} {fluid}"
        desc = desc or block
        for i in range(1, 11):
            v = f"{block}_earliest_visit_PC{i}"
            out.append(entry(
                category="Biologics", variable=v,
                description=f"Principal component {i} of the {desc} panel, earliest visit with data",
                original=f"{block}_* analytes", dataset=f"{block} to_merge",
                notes=(f"PCA over {int(r['features'])} analytes in {int(r['samples'])} participants "
                       f"(each participant's earliest visit carrying {block} data; "
                       f"{r['pct_samples_at_BL']}% of those are EVENT_ID=='BL'). "
                       f"Assays missing in >=20% of block samples dropped "
                       f"({int(r['features_dropped'])} dropped); residual gaps mean-imputed "
                       f"({r['pct_cells_imputed']}% of matrix cells); StandardScaler then "
                       f"PCA(n_components=10). PC1 explains {r['pc1_var']}%, PC1-10 "
                       f"{r['cum_var_pc1_10']}%. Scores are joined on PATNO so a participant's "
                       "value repeats across all of their visits. Proteomic analytes only — no "
                       "clinical, genetic or PLATE_ID column enters the PCA. PCs from different "
                       "blocks are different coordinate systems and must not be pooled.")))
    return out


BLOCK_DESC = {
    "olink_plasma": ("Olink Explore HT", "plasma"),
    "olink_csf": ("Olink Explore HT", "CSF"),
    "nulisa_cns_plasma": ("NULISA CNS Disease / Neuro 220", "plasma"),
    "nulisa_cns_csf": ("NULISA CNS Disease / Neuro 220", "CSF"),
    "nulisa_inf_plasma": ("NULISA Inflammation", "plasma"),
    "nulisa_inf_csf": ("NULISA Inflammation", "CSF"),
}


def build_harmonized_entries(fit: pd.DataFrame) -> list[dict]:
    """One entry per harmonized analyte, plus the per-block provenance stubs."""
    out = []
    for _, r in fit.iterrows():
        plat, fluid = BLOCK_DESC.get(r["block"], (r["block"], ""))
        slope_txt = (f"slope {r['applied_slope']:.3f} (bootstrap-supported)"
                     if r["slope_supported"] else
                     f"slope fixed at 1.000 (fitted {r['fitted_slope']:.3f}, "
                     f"95% CI [{r['slope_ci_lo']:.2f}, {r['slope_ci_hi']:.2f}] "
                     "spans the tolerance band, so no scale correction applied)")
        out.append(entry(
            category="Proteomics", variable=r["harmonized_column"],
            description=(f"{r['analyte']} in {fluid}, {plat}, harmonized across "
                         f"{r['reference_project']} and {r['mapped_project']}"),
            original=f"{r['col_reference']}, {r['col_mapped']}",
            dataset=f"{r['reference_project']} + {r['mapped_project']}",
            notes=(
                f"Harmonized block `{r['block']}`. Reference is {r['reference_project']} "
                f"(the project with more rows); {r['mapped_project']} is mapped onto its "
                "scale as slope*value + intercept, then used ONLY for rows the reference "
                f"does not cover — values are never averaged. Applied: {slope_txt}, "
                f"intercept {r['applied_intercept']:+.4f}.\n"
                f"Fitted on {r['n_matched_cells']} (EVENT_ID x COHORT) matched cells with "
                ">=10 samples per project, using medians and IQRs so cohort composition "
                "cannot drive the estimate. Values remain on the native "
                f"{r['metric']} scale and are NOT z-scored.\n"
                f"{r['pct_from_corrected']}% of this block's populated rows come from the "
                "corrected mapped project; the rest are uncorrected reference values. "
                "The project-specific columns are retained, so any result can be checked "
                "against the unpooled data. See the README for the validation suite and "
                "known limitations.")))
    for block in fit["block"].unique():
        sub = fit[fit.block == block].iloc[0]
        out.append(entry(
            category="Technical", variable=f"harmonized_{block}_src",
            description=f"Which project supplied this row's harmonized_{block} values",
            code="-", decode=f"{sub['reference_project']}; {sub['mapped_project']}_corrected",
            original=f"{sub['reference_project']}, {sub['mapped_project']}",
            dataset="Derived",
            notes=("Provenance for the harmonized block. Condition on this to restrict an "
                   "analysis to one source, or include it as a model term. Note a "
                   "participant's source can change between visits, which a rate-of-change "
                   "model can read as slope — restrict to non-switchers or carry the "
                   "source indicator for slope_* and tte_* outcomes.")))
    return out


def collection_era_entry() -> dict:
    return entry(
        category="Technical", variable="collection_era",
        description="Collection period of the visit, derived from visit_date",
        code="-", decode="pre_2020; 2020_plus",
        original="visit_date", dataset="PPMI_Curated_Data_Cut_Public_20260511",
        notes=("Added to support adjustment for a collection-era confound in Project 277. "
               "p277's post-2020 samples contain 2 healthy controls against 372 NSD+, and "
               "its whole panel is inflated as a result (lambda 2.486 against p314_CSF's "
               "1.113 under an identical model); its spike-in control p277_CSF_CTRL is "
               "significantly associated with NSD status at P=3.9e-4, which is impossible "
               "biologically.\n"
               "USE AS A CATEGORICAL COVARIATE, NOT AS A FILTER. This is a positivity "
               "violation: only 6 of 15 collection years carry both arms at n>=10, so in "
               "the rest the within-year group contrast is not estimable. A dummy per era "
               "absorbs those samples and they contribute nothing to the group "
               "coefficient — algebraically equivalent to restriction for the contrast, "
               "while retaining the samples for nuisance terms and adapting per contrast. "
               "A LINEAR year term is worse than useless (lambda 1.731 vs 1.459 "
               "categorical) because it extrapolates through the single-arm years.\n"
               "No samples are dropped: the post-2020 rows are unusable for an "
               "HC-referenced contrast but valid for within-PD or NSD-stage analyses. "
               "Residual inflation remains either way (lambda ~1.45 vs p314's 1.11), and "
               "era is only a proxy for assay batch, which p277 does not carry."))


def corrected_derivation_entries() -> dict[str, dict]:
    """Entries that must NOT be inherited from the previous release."""
    lowput_note = (
        "CORRECTED IN THIS RELEASE — values differ substantially from previous releases.\n"
        "Taken directly from MIA_LOWPUT_EXPECTED, which despite its name is already the "
        "observed/expected ratio: the curated cut defines it as "
        "Min(putamen_r_ref_cwm, putamen_l_ref_cwm) / (1.4474 - 0.003780*age_at_DATSCAN "
        "+ 0.2093*gender), MIAKAT pipeline, cerebral white matter reference region.\n"
        "Previous releases computed min(DATSCAN_PUTAMEN_L, R) / lowput_expected. But "
        "lowput_expected was itself documented as the 'age-/sex-expected lowest putamen "
        "ratio' — already observed/expected — so dividing by it a second time cancelled the "
        "imaging term and left a function of age and sex. Evidence in the previous release: "
        "lowput_ratio correlated 0.055 with the observed minimum putamen SBR; 0 of 4,312 "
        "values fell below 0.75 despite the entry stating '<0.75 indicates DaT positivity' "
        "(lowput_expected itself had 75.8% below 0.75).\n"
        "In this release lowput_ratio correlates 0.986 with the observed minimum putamen "
        "SBR and 62.5% of values fall below 0.75. Note the DaTscan series also changed: the "
        "curated public cut carries only the MIAKAT (MIA_*) reconstruction, not the DATSCAN_* "
        "series used previously, and the two are different quantifications (r ~= 0.89-0.90)."
    )
    return {
        "lowput_ratio": entry(
            category="DATSCAN", variable="lowput_ratio",
            description="Observed / age-and-sex-expected lowest-putamen SBR (MIAKAT pipeline), per visit",
            original="MIA_LOWPUT_EXPECTED", dataset="PPMI_Curated_Data_Cut_Public_20260511",
            notes=lowput_note),
        "slope_lowput_ratio": entry(
            category="DATSCAN", variable="slope_lowput_ratio",
            description="Per-participant OLS slope of lowput_ratio against YEAR",
            original="lowput_ratio, YEAR", dataset="Derived",
            notes=("Ordinary least squares slope across a participant's visits; one value per "
                   "PATNO repeated across all visits. Requires >=2 visits with a value and "
                   "non-zero spread in YEAR.\n"
                   "CORRECTED IN THIS RELEASE, inheriting the lowput_ratio fix above. In "
                   "previous releases this was near-constant across participants "
                   "(IQR -0.0158..-0.0141, sd 0.002), tracking the -0.01397/yr age coefficient "
                   "rather than any participant's dopaminergic decline. Prior results using it "
                   "should be re-read.")),
        "key": entry(
            category="ID", variable="key", derived="Yes",
            description="Merge key: PATNO + '_' + EVENT_ID",
            original="PATNO, EVENT_ID", dataset="Multiple",
            notes=("Built independently of MERGE_INDEX and asserted equal to it. Both key "
                   "components are normalised before concatenation — surrounding whitespace, "
                   "non-breaking and zero-width characters, float-formatted IDs ('3000.0') and "
                   "lowercase visit codes are repaired and reported; non-numeric PATNO, "
                   "malformed visit codes and duplicate keys abort the build.")),
    }


def main() -> None:
    # --dict-only regenerates the dictionary against the already-written dataset's
    # header, for iterating on entry text without rewriting a 1.2 GB table.
    dict_only = "--dict-only" in sys.argv
    log(f"=== Step 7: dataset + dictionary ===  timestamp {TIMESTAMP}"
        f"{'  [--dict-only]' if dict_only else ''}")

    f_data_existing = os.path.join(HERE, "Project_9004_Unified_Emerging_Biomarkers.tab")
    if dict_only:
        if not os.path.exists(f_data_existing):
            sys.exit("--dict-only needs an existing dataset to read the column order from")
        cols = pd.read_csv(f_data_existing, sep="\t", nrows=0).columns.tolist()
        df = pd.DataFrame(columns=cols)
        log(f"reusing existing dataset header: {len(cols):,} columns")
    else:
        merged = pd.read_csv(latest("merged-*.tab"), sep="\t", low_memory=False,
                             dtype={"key": str, "MERGE_INDEX": str, "PATNO": str, "EVENT_ID": str})
        derived = pd.read_csv(latest("derived-*.tab"), sep="\t", dtype={"key": str})
        pcs = pd.read_csv(latest("block_pcs-*.tab"), sep="\t", dtype={"key": str})
        log(f"merged {merged.shape}  derived {derived.shape}  pcs {pcs.shape}")

        for name, part in (("derived", derived), ("pcs", pcs)):
            assert part["key"].tolist() == merged["key"].tolist(), \
                f"{name} key order differs from merged"

        harm = pd.read_csv(latest("harmonized-*.tab"), sep="\t", low_memory=False,
                           dtype={"key": str, "collection_era": str})
        hpcs = pd.read_csv(latest("harmonized_pcs-*.tab"), sep="\t", dtype={"key": str})
        log(f"harmonized {harm.shape}  harmonized_pcs {hpcs.shape}")
        for nm, part in (("harmonized", harm), ("harmonized_pcs", hpcs)):
            assert part["key"].tolist() == merged["key"].tolist(), \
                f"{nm} key order differs from merged"
        df = pd.concat([merged, derived.drop(columns=["key"]),
                        pcs.drop(columns=["key"]),
                        harm.drop(columns=["key"]), hpcs.drop(columns=["key"])], axis=1)
        assert len(df) == len(merged), "row count changed during assembly"
        assert not df.columns.duplicated().any(), "duplicate column names in final dataset"
        log(f"final dataset: {df.shape[0]:,} rows x {df.shape[1]:,} columns")

    # --- dictionary --------------------------------------------------------
    scaffold_dict = pd.read_csv(latest("scaffold_dict-*.tab"), sep="\t")
    prev = pd.read_csv(PREV_DICT, sep="\t")
    log(f"scaffold dict {len(scaffold_dict)}  previous release dict {len(prev)}")

    lookup: dict[str, dict] = {}
    for _, r in prev.iterrows():                      # lowest priority
        lookup[r["Variable"]] = r[DICT_COLS].to_dict()
    for _, r in scaffold_dict.iterrows():             # clinical is authoritative
        lookup[r["Variable"]] = r[DICT_COLS].to_dict()

    p9001_cols = [c for c in df.columns if c.startswith("p9001_")]
    for e in build_p9001_entries(p9001_cols):
        lookup[e["Variable"]] = e

    variance = pd.read_csv(latest("pc_variance-*.tab"), sep="\t")
    for e in build_pc_entries(variance):
        lookup[e["Variable"]] = e

    fit = pd.read_csv(latest("harmonization_fit-*.tab"), sep="\t")
    for e in build_harmonized_entries(fit):
        lookup[e["Variable"]] = e
    hvar = pd.read_csv(latest("harmonized_pc_variance-*.tab"), sep="\t")
    for e in build_pc_entries(hvar):
        lookup[e["Variable"]] = e
    lookup["collection_era"] = collection_era_entry()

    for v, e in corrected_derivation_entries().items():
        lookup[v] = e

    missing = [c for c in df.columns if c not in lookup]
    if missing:
        log(f"  {len(missing)} column(s) with no dictionary entry: {missing[:20]}")
        sys.exit("cannot write a 1:1 dictionary")

    dd = pd.DataFrame([lookup[c] for c in df.columns], columns=DICT_COLS)
    dd["Derivation Notes"] = dd["Derivation Notes"].fillna("")

    assert len(dd) == df.shape[1], "dictionary length != dataset column count"
    assert dd["Variable"].tolist() == list(df.columns), "dictionary order != dataset order"
    assert dd["Variable"].is_unique, "duplicate Variable in dictionary"
    log(f"dictionary is 1:1 with the dataset ({len(dd):,} entries)")
    log("\ncategory breakdown:")
    for k, v in dd["Category"].value_counts().items():
        log(f"  {k:16s} {v:,}")

    # --- write -------------------------------------------------------------
    f_data = os.path.join(HERE, "Project_9004_Unified_Emerging_Biomarkers.tab")
    f_dict = os.path.join(HERE, "Project_9004_Data_Dictionary.tab")
    f_xlsx = os.path.join(HERE, "Project_9004_Data_Dictionary.xlsx")

    if not dict_only:
        log(f"\nwriting {os.path.basename(f_data)} ...")
        df.to_csv(f_data, sep="\t", index=False)
        log(f"wrote {os.path.basename(f_data)}  ({os.path.getsize(f_data) / 1e9:.2f} GB)")
    else:
        log("\n--dict-only: dataset left untouched")
    dd.to_csv(f_dict, sep="\t", index=False)
    write_xlsx(dd, f_xlsx)

    log(f"wrote {os.path.basename(f_dict)}  ({len(dd):,} entries)")
    log(f"wrote {os.path.basename(f_xlsx)}")

    n_nl = int(dd.map(lambda x: isinstance(x, str) and "\n" in x).sum().sum())
    log(f"\n{n_nl} dictionary cells contain embedded newlines — this is why the .xlsx "
        "exists alongside the .tab (Excel's text import splits them across rows).")


def write_xlsx(dd: pd.DataFrame, path: str) -> None:
    """One row per entry, with newlines living inside cells rather than breaking rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"
    ws.append(list(dd.columns))
    for row in dd.itertuples(index=False):
        ws.append(["" if pd.isna(v) else str(v) for v in row])

    for c in range(1, len(dd.columns) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {"Category": 16, "Variable": 42, "Description": 62, "Code": 14, "Decode": 26,
              "Derived Variable": 10, "Original Variable(s)": 30,
              "Original Dataset(s)": 26, "Derivation Notes": 90}
    for i, col in enumerate(dd.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 20)
        for cell in ws[get_column_letter(i)]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
